"""
Extract the 49 active Colorado birthing hospitals from the CHA Master Hospital
List XLSX into hospitals_master_2026.json — the JSON read by both the backend
seed and the PM data-entry workbook generator.

Source: Hospitals.xlsx (CHA Master Hospital List sheet)
Output: data/hospitals_master_2026.json

Usage:
    python3 data/extract_master_hospitals.py <path-to-Hospitals.xlsx>

If CHA updates the master list, re-run this script to refresh
hospitals_master_2026.json, then re-run:
    npm run db:seed          # picks up new hospitals and metadata
    npm run db:enroll        # picks up updated participation flags
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

DEFAULT_OUT = Path(__file__).parent / "hospitals_master_2026.json"


def yes(v):
    return str(v).strip().lower() in ("yes", "true", "y", "1")


def s(v):
    if v is None:
        return None
    out = str(v).strip()
    return out or None


def i(v):
    if v is None:
        return None
    try:
        if isinstance(v, str) and not re.fullmatch(r"-?\d+(\.0+)?", v):
            return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def extract(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Master Hospital List"]
    headers = [c.value for c in ws[1]]

    def col(name):
        return headers.index(name) + 1

    hospitals = []
    for r in range(2, ws.max_row + 1):
        name = s(ws.cell(row=r, column=col("Hospital_Name_CHA")).value)
        if not name:
            continue
        if not yes(ws.cell(row=r, column=col("Active Birthing Hospital Current Year")).value):
            continue
        hospitals.append(
            {
                "chaHospitalId": s(ws.cell(row=r, column=col("CHA_Hospital_ID")).value),
                "name": name,
                "cdpheName": s(ws.cell(row=r, column=col("Hospital_Name_CDPHE")).value),
                "tableauNickname": s(ws.cell(row=r, column=col("Tableau Nickname (CPCQC)")).value),
                "system": s(ws.cell(row=r, column=col("Hospital System")).value),
                "address": s(ws.cell(row=r, column=col("Address")).value),
                "city": s(ws.cell(row=r, column=col("City")).value),
                "state": "CO",
                "postalCode": s(ws.cell(row=r, column=col("Zipcode")).value),
                "county": s(ws.cell(row=r, column=col("County")).value),
                "cdpheId": s(ws.cell(row=r, column=col("ID_CDPHE")).value),
                "aimId": s(ws.cell(row=r, column=col("ID_AIM")).value),
                "nicuLevel": s(ws.cell(row=r, column=col("NICU Level")).value),
                "urbanicity": s(ws.cell(row=r, column=col("Urban/Rural")).value),
                "birthVolume2025": i(ws.cell(row=r, column=col("Birth Volume 2025")).value),
                "rae": i(ws.cell(row=r, column=col("RAE")).value),
                "hsr": i(ws.cell(row=r, column=col("HSR")).value),
                "participation": {
                    "SOAR": {
                        "participating": yes(ws.cell(row=r, column=col("Participating SOAR Current Year")).value),
                        "status2026": s(ws.cell(row=r, column=col("SOAR Status 2026")).value),
                    },
                    "TTT": {
                        "participating": yes(ws.cell(row=r, column=col("Participating TtT Current Year")).value),
                    },
                    "SPARK": {
                        "participating": yes(ws.cell(row=r, column=col("Participating SPARK Current Year")).value),
                    },
                    "NEST": {
                        "participating": yes(ws.cell(row=r, column=col("Participating NEST Current Year")).value),
                    },
                },
            }
        )

    return {
        "asOf": "2026",
        "source": "CHA Master Hospital List (Hospitals.xlsx, Master Hospital List sheet)",
        "taxonomy": "CHA",
        "count": len(hospitals),
        "hospitals": hospitals,
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 data/extract_master_hospitals.py <path-to-Hospitals.xlsx>")
        sys.exit(1)
    src = Path(sys.argv[1])
    if not src.exists():
        print(f"Source file not found: {src}")
        sys.exit(1)
    data = extract(src)
    DEFAULT_OUT.write_text(json.dumps(data, indent=2))
    print(f"Wrote {DEFAULT_OUT} — {data['count']} hospitals")


if __name__ == "__main__":
    main()
